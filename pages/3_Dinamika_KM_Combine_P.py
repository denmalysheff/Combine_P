import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Dinamik_kilometr", layout="wide")

# Заголовок и описание
st.title("📊 Dinamik_kilometr")
st.write("Загрузите два Excel-файла для сравнения динамики оценок.")

def load_and_filter_data(file):
    try:
        df = pd.read_excel(file, sheet_name="Оценка КМ")
        df.columns = df.columns.str.strip()
        # Твои фильтры из исходного кода
        df = df[df["КОДНАПР"].isin([24602, 24607, 24701, 24603])]
        return df[["КОДНАПР", "ПУТЬ", "ПД", "KM", "ОЦЕНКА"]]
    except Exception as e:
        st.error(f"Ошибка при чтении листа 'Оценка КМ': {e}")
        return None

def analyze_changes(df1, df2):
    grade_map = {5: "отлично", 4: "хорошо", 3: "удовлетворительно", 2: "неудовлетворительно"}
    merged_df = df1.merge(df2, on=["КОДНАПР", "ПУТЬ", "ПД", "KM"], suffixes=("_old", "_new"))
    merged_df["Переход"] = merged_df.apply(
        lambda row: f"{grade_map.get(row['ОЦЕНКА_old'], 'неизвестно')} -> {grade_map.get(row['ОЦЕНКА_new'], 'неизвестно')}",
        axis=1
    )
    return merged_df, merged_df["Переход"].value_counts().to_dict()

def apply_excel_formatting(output_buffer):
    """Применяет твою логику раскраски строк в Excel"""
    wb = load_workbook(output_buffer)
    fills = {
        "синий": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "оранжевый": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "зеленый": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "красный": PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid"),
    }
    grade_map = {"неудовлетворительно": 2, "удовлетворительно": 3, "хорошо": 4, "отлично": 5}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.max_row < 2: continue
        header = [cell.value for cell in ws[1]]
        if "Переход" not in header: continue
        
        transition_idx = header.index("Переход")
        for row in ws.iter_rows(min_row=2):
            cell_value = str(row[transition_idx].value)
            fill = None
            if " -> " in cell_value:
                old_t, new_t = cell_value.split(" -> ")
                old_s, new_s = grade_map.get(old_t), grade_map.get(new_t)
                if old_s and new_s:
                    if old_s == 2 and new_s == 2: fill = fills["красный"]
                    elif old_s == new_s: fill = fills["синий"]
                    elif old_s < new_s: fill = fills["зеленый"]
                    elif old_s > new_s: fill = fills["оранжевый"]
            
            if fill:
                for cell in row: cell.fill = fill
    
    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()

# Интерфейс загрузки
col1, col2 = st.columns(2)
with col1:
    file1 = st.file_uploader("Первый файл (Excel)", type=["xlsx"])
with col2:
    file2 = st.file_uploader("Второй файл (Excel)", type=["xlsx"])

if file1 and file2:
    df1 = load_and_filter_data(file1)
    df2 = load_and_filter_data(file2)

    if df1 is not None and df2 is not None:
        if st.button("🚀 Обработать и подготовить отчет", use_container_width=True):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Группировка по ПД
                pd_values = sorted(df1["ПД"].unique())
                for pd_val in pd_values:
                    d1_sub = df1[df1["ПД"] == pd_val]
                    d2_sub = df2[df2["ПД"] == pd_val]
                    if not d1_sub.empty and not d2_sub.empty:
                        c_df, summ = analyze_changes(d1_sub, d2_sub)
                        c_df.to_excel(writer, sheet_name=f"ПД-{pd_val}_детально", index=False)
                        pd.DataFrame(summ.items(), columns=["Переход", "Кол-во"]).to_excel(writer, sheet_name=f"ПД-{pd_val}_сводка", index=False)

                # Группировка по ПЧУ/ПЧЗ
                groups = {
                    "ПЧУ-1": [1, 2, 3], "ПЧУ-2": [4, 5, 12], 
                    "ПЧЗ-ЮГ": [1, 2, 3, 4, 5, 12],
                    "ПЧЗ-ЗАПАД": [6, 7, 8, 9, 10, 11, 13, 14, 15], 
                    "ПЧ": list(range(1, 16))
                }
                for name, pds in groups.items():
                    d1_g = df1[df1["ПД"].isin(pds)]
                    d2_g = df2[df2["ПД"].isin(pds)]
                    if not d1_g.empty and not d2_g.empty:
                        c_df, summ = analyze_changes(d1_g, d2_g)
                        c_df.to_excel(writer, sheet_name=f"{name}_детально", index=False)
                        pd.DataFrame(summ.items(), columns=["Переход", "Кол-во"]).to_excel(writer, sheet_name=f"{name}_сводка", index=False)

            # Форматирование и выдача файла
            processed_data = apply_excel_formatting(output)
            st.success("Готово! Нажмите кнопку ниже, чтобы скачать файл.")
            st.download_button(
                label="📥 Скачать результат",
                data=processed_data,
                file_name="Dinamika_km_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )