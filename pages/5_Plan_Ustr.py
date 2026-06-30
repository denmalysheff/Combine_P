import pandas as pd
import streamlit as st
import io
from datetime import datetime
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def normalize_column_name(col):
    """Нормализует имя столбца или вкладки для защиты от ошибок раскладки (РУС/ENG) и пробелов"""
    if not isinstance(col, str):
        return col
    col = col.strip().upper().replace(" ", "").replace("_", "")
    replacements = {
        'K': 'К', 'M': 'М', 'P': 'Р', 'A': 'А', 'C': 'С', 
        'O': 'О', 'T': 'Т', 'B': 'В', 'E': 'Е', 'H': 'Н', 'X': 'Х'
    }
    for eng, rus in replacements.items():
        col = col.replace(eng, rus)
    return col

def find_column(df, target_name):
    """Ищет столбец в листе Excel по его нормализованному имени"""
    norm_target = normalize_column_name(target_name)
    for col in df.columns:
        if normalize_column_name(col) == norm_target:
            return col
    return None

def find_sheet_smart(excel_file, target_name):
    """Умный поиск имени вкладки в файле Excel без привязки к регистру, пробелам и раскладке"""
    norm_target = normalize_column_name(target_name)
    for sheet in excel_file.sheet_names:
        if normalize_column_name(sheet) == norm_target:
            return sheet
    return None

def safe_int(value, default=0):
    """Безопасно преобразует значение в int, защищая от прочерков '-', nan и текста"""
    if pd.isna(value):
        return default
    val_str = str(value).strip()
    if val_str in ['-', '', 'nan', 'None', '.0']:
        return default
    try:
        return int(float(val_str))
    except ValueError:
        return default

def process_track_data(uploaded_file, inc_bridge_objects=False, inc_is=False,
                       inc_dnprof=False, inc_pru=False, inc_anp=False, inc_zaz=False, inc_rshk=False):
    try:
        xl = pd.ExcelFile(uploaded_file)
        
        sheet_assessment_name = find_sheet_smart(xl, 'ОЦЕНКАКМ')
        sheet_defects_name = find_sheet_smart(xl, 'ОТСТУПЛЕНИЯ')
        
        if not sheet_assessment_name:
            raise ValueError("В исходном файле не найдена вкладка 'Оценка КМ'.")
        if not sheet_defects_name:
            raise ValueError("В исходном файле не найдена вкладка 'Отступления'.")
            
        df_assessment = xl.parse(sheet_assessment_name)
        df_defects = xl.parse(sheet_defects_name)
    except Exception as e:
        raise ValueError(f"Ошибка при чтении файла Excel: {e}")

    # --- Сбор и нормализация столбцов 'Оценка КМ' ---
    col_kodnapr = find_column(df_assessment, 'КОДНАПР')
    col_path_assess = find_column(df_assessment, 'ПУТЬ')
    col_km_assess = find_column(df_assessment, 'КМ')
    col_pd_assess = find_column(df_assessment, 'ПД')
    col_rating = find_column(df_assessment, 'ОЦЕНКА')
    col_score_assess = find_column(df_assessment, 'БАЛЛ')
    col_lim_pass = find_column(df_assessment, 'СК_ОГР_ПАСС')
    col_lim_gruz = find_column(df_assessment, 'СК_ОГР_ГРУЗ')

    if not col_kodnapr or not col_km_assess:
        raise KeyError("В листе 'Оценка КМ' не найдены обязательные столбцы КОДНАПР или КМ.")

    # Фильтрация по списку направлений ПЧ-22
    target_directions = [24701, 24602, 24603]
    df_assess_filtered = df_assessment[df_assessment[col_kodnapr].isin(target_directions)].copy()
    if df_assess_filtered.empty:
        raise ValueError("В листе 'Оценка КМ' не найдено записей для направлений 24701, 24602, 24603")

    df_assess_filtered['MATCH_ПУТЬ'] = df_assess_filtered[col_path_assess].apply(safe_int)
    df_assess_filtered['MATCH_КМ'] = df_assess_filtered[col_km_assess].apply(safe_int)

    def make_speed_limit(row):
        p = str(row[col_lim_pass]).split('.')[0].strip() if col_lim_pass and pd.notna(row[col_lim_pass]) else "-"
        g = str(row[col_lim_gruz]).split('.')[0].strip() if col_lim_gruz and pd.notna(row[col_lim_gruz]) else "-"
        if p in ['', 'nan', '-']: p = "-"
        if g in ['', 'nan', '-']: g = "-"
        if p == "-" and g == "-": return ""
        return f"{p}/{g}"

    df_assess_filtered['ОГРАНИЧЕНИЕ_СКОРОСТИ'] = df_assess_filtered.apply(make_speed_limit, axis=1)

    # --- Сбор и нормализация столбцов 'Отступления' ---
    col_path_def = find_column(df_defects, 'ПУТЬ')
    col_km_def = find_column(df_defects, 'КМ М') or find_column(df_defects, 'КМ')
    col_meter = find_column(df_defects, 'МЕТР') or find_column(df_defects, 'М')
    col_defect = find_column(df_defects, 'ОТСТУПЛЕНИЕ')
    col_degree = find_column(df_defects, 'СТЕПЕНЬ')
    col_ampl = find_column(df_defects, 'АМПЛИТУДА')
    col_len = find_column(df_defects, 'ДЛИНА')
    col_is = find_column(df_defects, 'ИС')
    col_str = find_column(df_defects, 'СТРЕЛКА')
    col_obk = find_column(df_defects, 'ОБК')
    col_most = find_column(df_defects, 'МОСТ')
    col_pr = find_column(df_defects, 'PR_PREDUPR')
    col_score_def = find_column(df_defects, 'БАЛЛ')
    col_def_lim_p = find_column(df_defects, 'СК_ОГР_ПАСС') or find_column(df_defects, 'ДОП_ПАСС')

    if not col_km_def or not col_degree:
        raise KeyError("В листе 'Отступления' не найдены столбцы КМ или СТЕПЕНЬ.")

    df_defects['MATCH_ПУТЬ'] = df_defects[col_path_def].apply(safe_int)
    df_defects['MATCH_КМ'] = df_defects[col_km_def].apply(safe_int)
    df_defects['INT_СТЕПЕНЬ'] = df_defects[col_degree].apply(safe_int)

    if col_str:
        df_defects['FILTER_СТРЕЛКА'] = df_defects[col_str].apply(safe_int)
        df_defects = df_defects[df_defects['FILTER_СТРЕЛКА'] == 0]

    # Маркировка типов неисправностей для фильтрации и штучного подсчета
    if col_defect:
        df_defects['TMP_DEF_UPPER'] = df_defects[col_defect].astype(str).str.strip().str.upper()
    else:
        df_defects['TMP_DEF_UPPER'] = ""

    df_defects['IS_DNPROF_VAL'] = df_defects['TMP_DEF_UPPER'] == 'ДНПРОФ'
    df_defects['IS_PRU_VAL'] = df_defects['TMP_DEF_UPPER'] == 'ПРУ'
    df_defects['IS_ANP_VAL'] = df_defects['TMP_DEF_UPPER'] == 'АНП'
    df_defects['IS_ZAZ_VAL'] = df_defects['TMP_DEF_UPPER'] == 'ЗАЗ'
    df_defects['IS_RSHK_VAL'] = df_defects['TMP_DEF_UPPER'] == 'РШК'

    df_defects['IS_SPECIAL_PARAM'] = (
        df_defects['IS_DNPROF_VAL'] | df_defects['IS_PRU_VAL'] | 
        df_defects['IS_ANP_VAL'] | df_defects['IS_ZAZ_VAL'] | df_defects['IS_RSHK_VAL']
    )

    # Фильтр по чекбоксам: оставляем только то, что пользователь ВКЛЮЧИЛ
    def keep_by_checkbox(row):
        if not row['IS_SPECIAL_PARAM']:
            return True 
        if row['IS_DNPROF_VAL'] and inc_dnprof: return True
        if row['IS_PRU_VAL'] and inc_pru: return True
        if row['IS_ANP_VAL'] and inc_anp: return True
        if row['IS_ZAZ_VAL'] and inc_zaz: return True
        if row['IS_RSHK_VAL'] and inc_rshk: return True
        return False

    df_defects = df_defects[df_defects.apply(keep_by_checkbox, axis=1)].copy()

    # --- ПРОВЕРКА ОГРАНИЧЕНИЙ ДЛЯ 2 СТЕПЕНИ (2к3ст) ---
    def check_is_2_pr(row):
        if row['IS_SPECIAL_PARAM']: return False 
        if safe_int(row['INT_СТЕПЕНЬ']) != 2: return False
        if col_pr:
            pr_val = str(row[col_pr]).strip().upper()
            if pr_val not in ['0', '0.0', '', 'NAN', 'NONE', '-']: return True
        if col_def_lim_p and pd.notna(row[col_def_lim_p]):
            try:
                spd_p = float(row[col_def_lim_p])
                if 0 < spd_p < 140: return True
            except ValueError: pass
        return False

    df_defects['IS_2_PR'] = df_defects.apply(check_is_2_pr, axis=1)
    df_defects['IS_2_REGULAR'] = (df_defects['INT_СТЕПЕНЬ'] == 2) & (~df_defects['IS_2_PR']) & (~df_defects['IS_SPECIAL_PARAM'])
    df_defects['IS_3'] = (df_defects['INT_СТЕПЕНЬ'] == 3) & (~df_defects['IS_SPECIAL_PARAM'])
    df_defects['IS_4'] = (df_defects['INT_СТЕПЕНЬ'] == 4) & (~df_defects['IS_SPECIAL_PARAM'])

    # --- ДОПОЛНИТЕЛЬНЫЕ ФИЛЬТРЫ ВЫБОРКИ ---
    if col_defect:
        # 1. Включить просадки на ИС
        if col_is and inc_is:
            df_defects['IS_PROSADKA_IS'] = (
                df_defects['TMP_DEF_UPPER'].isin(['ПР.Л', 'ПР.П', 'ПР Л', 'ПР П']) & 
                (df_defects['INT_СТЕПЕНЬ'] > 1) & 
                (df_defects[col_is].apply(safe_int) == 1)
            )
        else:
            df_defects['IS_PROSADKA_IS'] = False
            
        # 2. Критическое уширение колеи от 1545 мм
        if col_ampl:
            df_defects['IS_USH_CRITICAL'] = (
                df_defects['TMP_DEF_UPPER'].isin(['УШ', 'УШ.']) & 
                (df_defects[col_ampl].apply(safe_int) >= 1545)
            )
        else:
            df_defects['IS_USH_CRITICAL'] = False

        # 3. Включить просадки > 20 мм на мостах или объектах
        if inc_bridge_objects and col_ampl and (col_most or col_obk):
            is_bridge = df_defects[col_most].apply(safe_int) == 1 if col_most else False
            is_object = df_defects[col_obk].apply(safe_int) == 1 if col_obk else False
            df_defects['IS_BRIDGE_OBJECT_CRITICAL'] = (
                df_defects['TMP_DEF_UPPER'].isin(['ПР.Л', 'ПР.П', 'ПР Л', 'ПР П', 'П', 'П.']) &
                (df_defects[col_ampl].apply(safe_int) > 20) & (is_bridge | is_object)
            )
        else:
            df_defects['IS_BRIDGE_OBJECT_CRITICAL'] = False
    else:
        df_defects['IS_PROSADKA_IS'] = False
        df_defects['IS_USH_CRITICAL'] = False
        df_defects['IS_BRIDGE_OBJECT_CRITICAL'] = False

    df_valid_defects = df_defects[
        df_defects['INT_СТЕПЕНЬ'].isin([2, 3, 4]) | 
        df_defects['IS_PROSADKA_IS'] | 
        df_defects['IS_USH_CRITICAL'] |
        df_defects['IS_BRIDGE_OBJECT_CRITICAL'] |
        df_defects['IS_SPECIAL_PARAM']
    ].copy()

    def aggregate_km_defects(group):
        count_2 = group['IS_2_REGULAR'].sum()
        count_2_3 = group['IS_2_PR'].sum()
        count_3 = group['IS_3'].sum()
        count_4 = group['IS_4'].sum()
        
        dnprof_c = group['IS_DNPROF_VAL'].sum()
        pru_c = group['IS_PRU_VAL'].sum()
        anp_c = group['IS_ANP_VAL'].sum()
        zaz_c = group['IS_ZAZ_VAL'].sum()
        rshk_c = group['IS_RSHK_VAL'].sum()
        
        text_rows = group[
            (group['IS_3'] | group['IS_4'] | group['IS_2_PR'] | 
             group['IS_PROSADKA_IS'] | group['IS_USH_CRITICAL'] | group['IS_BRIDGE_OBJECT_CRITICAL']) & 
            (~group['IS_SPECIAL_PARAM'])
        ]
        
        if col_meter and col_defect:
            text_rows = text_rows.drop_duplicates(subset=[col_meter, col_defect])
        
        list_desc = []
        for _, row in text_rows.iterrows():
            m_val = str(safe_int(row[col_meter])) if col_meter else "0"
            def_type = str(row[col_defect]).strip() if col_defect else "ОТСТ"
            deg_str = "2к3ст" if row['IS_2_PR'] else f"{safe_int(row['INT_СТЕПЕНЬ'])}ст"
            ampl = str(safe_int(row[col_ampl])) if col_ampl else "0"
            length = str(safe_int(row[col_len])) if col_len else "0"
            def_score = str(safe_int(row[col_score_def])) if col_score_def else "0"
            
            tags = []
            if col_obk and pd.notna(row[col_obk]) and str(row[col_obk]).strip() not in ['0', '0.0', '', '-']: tags.append("обк")
            if col_most and pd.notna(row[col_most]) and str(row[col_most]).strip() not in ['0', '0.0', '', '-']: tags.append("м")
            if col_is and pd.notna(row[col_is]) and str(row[col_is]).strip() not in ['0', '0.0', '', '-']: tags.append("ис")
            
            tag_str = " " + " ".join(tags) if tags else ""
            desc = f"{m_val}-{def_type} {deg_str} {ampl}/{length}{tag_str.lower()} {def_score}б."
            list_desc.append(desc)
            
        final_text = ", ".join(list_desc) if list_desc else ""
        
        counts_list = []
        if dnprof_c > 0: counts_list.append(f"Днпроф - {dnprof_c} шт.")
        if pru_c > 0: counts_list.append(f"ПрУ - {pru_c} шт.")
        if anp_c > 0: counts_list.append(f"Анп - {anp_c} шт.")
        if zaz_c > 0: counts_list.append(f"Заз - {zaz_c} шт.")
        if rshk_c > 0: counts_list.append(f"РШК - {rshk_c} шт.")
        
        if counts_list:
            counts_str = ", ".join(counts_list)
            if final_text:
                final_text += f", {counts_str}"
            else:
                final_text = counts_str
        
        return pd.Series({
            'КОЛ_ВО_2': count_2,
            'КОЛ_ВО_2_3': count_2_3,
            'КОЛ_ВО_3': count_3,
            'КОЛ_ВО_4': count_4,
            'ПЕРЕЧЕНЬ_ОТСТУПЛЕНИЙ': final_text
        })

    if not df_valid_defects.empty:
        df_aggregated = df_valid_defects.groupby(['MATCH_ПУТЬ', 'MATCH_КМ']).apply(aggregate_km_defects, include_groups=False).reset_index()
        result_df = pd.merge(df_assess_filtered, df_aggregated, on=['MATCH_ПУТЬ', 'MATCH_КМ'], how='left')
    else:
        result_df = df_assess_filtered.copy()
        result_df['КОЛ_ВО_2'] = 0
        result_df['КОЛ_ВО_2_3'] = 0
        result_df['КОЛ_ВО_3'] = 0
        result_df['КОЛ_ВО_4'] = 0
        result_df['ПЕРЕЧЕНЬ_ОТСТУПЛЕНИЙ'] = ""

    for col in ['КОЛ_ВО_2', 'КОЛ_ВО_2_3', 'КОЛ_ВО_3', 'КОЛ_ВО_4']:
        result_df[col] = result_df[col].fillna(0).astype(int)
        result_df[col] = result_df[col].apply(lambda x: "" if x == 0 else x)
        
    result_df['ПЕРЕЧЕНЬ_ОТСТУПЛЕНИЙ'] = result_df['ПЕРЕЧЕНЬ_ОТСТУПЛЕНИЙ'].fillna("")

    # Формируем структуру без столбца исполнителя
    output_data = {
        'Направление': result_df[col_kodnapr],
        'ПД': result_df[col_pd_assess] if col_pd_assess else "",
        'Путь': result_df['MATCH_ПУТЬ'],
        'Км': result_df['MATCH_КМ'],
        'Оценка': result_df[col_rating],
        'Балл': result_df[col_score_assess],
        '2 ст': result_df['КОЛ_ВО_2'],
        '2к3ст': result_df['КОЛ_ВО_2_3'],
        '3 ст': result_df['КОЛ_ВО_3'],
        '4 ст': result_df['КОЛ_ВО_4'],
        'Огр. скорости': result_df['ОГРАНИЧЕНИЕ_СКОРОСТИ'],
        'Перечень отступлений': result_df['ПЕРЕЧЕНЬ_ОТСТУПЛЕНИЙ']
    }
    final_df = pd.DataFrame(output_data)

    # Добавляем 10 пустых столбцов для графика (без дат в названиях)
    for i in range(1, 11):
        final_df[f"График_{i}"] = ""

    def pd_sheet_sort_key(name):
        digits = ''.join(c for c in str(name) if c.isdigit())
        return int(digits) if digits else 999

    final_df['Околоток'] = final_df['Околоток'].astype(str).str.strip()
    unique_pds = [x for x in final_df['Околоток'].unique() if x not in ['', 'nan', 'None']]
    unique_pds_sorted = sorted(unique_pds, key=pd_sheet_sort_key)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for pd_name in unique_pds_sorted:
            df_pd = final_df[final_df['Околоток'] == pd_name].copy()
            df_pd.sort_values(by=['Направление', 'Путь', 'Км'], ascending=[True, True, True], inplace=True)
            
            sheet_title = f"ПД-{pd_name}" if not str(pd_name).startswith('ПД') else str(pd_name)
            df_pd.to_excel(writer, sheet_name=sheet_title, index=False, startrow=3)
            
            workbook = writer.book
            worksheet = workbook[sheet_title]
            
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
            
            # Настройки стилей текста
            align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')
            
            font_title = Font(name='Arial', size=16, bold=True)
            font_header = Font(name='Arial', size=10, bold=False)  # Нежирная аккуратная шапка таблицы
            font_bold_cell = Font(name='Arial', size=11, bold=True)
            font_normal_cell = Font(name='Arial', size=11, bold=False)
            
            # Настройки рамок (Сетка)
            thin_side = Side(border_style="thin", color="000000")
            thick_side = Side(border_style="medium", color="000000")
            border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # Крупный заголовок листа (строки 1-2)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=worksheet.max_column)
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = f"ПЛАН УСТРАНЕНИЯ ОТСТУПЛЕНИЙ {sheet_title.upper()}"
            title_cell.font = font_title
            title_cell.alignment = align_center
            
            worksheet.row_dimensions[4].height = 35  # Высота шапки таблицы
            
            # Оформление заголовков таблицы (строка 4)
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=4, column=col_idx)
                cell.alignment = align_header
                cell.font = font_header
                cell.border = border_cell
                
                # Затираем текст "График_Х" для календарной сетки, оставляя ячейки пустыми под ручную запись
                if col_idx > 12:
                    cell.value = ""
            
            # Оформление строк данных (начиная со строки 5)
            for row_idx in range(5, worksheet.max_row + 1):
                worksheet.row_dimensions[row_idx].height = 24
                rating_value = str(worksheet.cell(row=row_idx, column=5).value).strip()
                is_rating_2 = (rating_value == '2' or rating_value == '2.0')
                
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = border_cell
                    
                    if col_idx == 12:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center
                        
                    cell.font = font_bold_cell if is_rating_2 else font_normal_cell

            # Применение КРУПНОЙ внешней рамки на весь печатный блок таблицы (строки 4..max_row, колонки 1..max_col)
            for r in range(4, worksheet.max_row + 1):
                for c in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=r, column=c)
                    # Сборка составной рамки для границ контура
                    l = thick_side if c == 1 else cell.border.left
                    rt = thick_side if c == worksheet.max_column else cell.border.right
                    t = thick_side if r == 4 else cell.border.top
                    b = thick_side if r == worksheet.max_row else cell.border.bottom
                    cell.border = Border(left=l, right=rt, top=t, bottom=b)

            # Безопасная настройка ширины столбцов под А3
            for col_idx in range(1, worksheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                
                if col_idx in [1, 2, 5, 6]:
                    worksheet.column_dimensions[col_letter].width = 13
                elif col_idx in [3, 4, 7, 8, 9, 10]:
                    worksheet.column_dimensions[col_letter].width = 9
                elif col_idx == 11:  # Скорость
                    worksheet.column_dimensions[col_letter].width = 14
                elif col_idx == 12:  # Большое описание неисправностей километра
                    worksheet.column_dimensions[col_letter].width = 75
                else:  # Столбцы под ручные даты
                    worksheet.column_dimensions[col_letter].width = 7
                    
    processed_data = output.getvalue()
    return processed_data

# --- Веб-Интерфейс Streamlit ---
st.set_page_config(page_title="ЧП-22 Планирование", layout="wide")
st.title("🚂 ЧП-22 Планирование")
st.subheader("Аналитика вагонных данных и автоматическое распределение План-Графика")

uploaded_file = st.file_uploader("Выберите исходный Excel-файл с вагона", type=["xlsx", "xls"])

st.markdown("### 🛠️ Настройки включаемых в анализ параметров путеизмерителя (по умолчанию отключены):")
col1, col2 = st.columns(2)

with col1:
    inc_bridge_objects = st.checkbox("Включить выборку просадок > 20 мм на МОСТАХ и ОБЪЕКТАХ (МОСТ=1 / ОБК=1)", value=False)
    inc_is = st.checkbox("Включить просадки на изолированных стыках (ИС)", value=False)
    inc_dnprof = st.checkbox("Включить продольный профиль (ДНПРОФ)", value=False)
    inc_pru = st.checkbox("Включить просадку уровня на сопряжении (ПРУ)", value=False)

with col2:
    inc_anp = st.checkbox("Включить уклон отвода возвышения (АНП)", value=False)
    inc_zaz = st.checkbox("Включить стыковые зазоры (ЗАЗ)", value=False)
    inc_rshk = st.checkbox("Включить регулировку ширины колеи (РШК)", value=False)

if uploaded_file is not None:
    if st.button("Сформировать План-График ЧП-22", type="primary"):
        with st.spinner("⏳ Выполняется формирование строгих печатных форм для А3..."):
            try:
                excel_data = process_track_data(
                    uploaded_file, 
                    inc_bridge_objects=inc_bridge_objects,
                    inc_is=inc_is,
                    inc_dnprof=inc_dnprof,
                    inc_pru=inc_pru,
                    inc_anp=inc_anp,
                    inc_zaz=inc_zaz,
                    inc_rshk=inc_rshk
                )
                
                current_time = datetime.now().strftime("%d-%m-%Y_%H-%M")
                file_title = f"PCH22_Plan_Grafik_{current_time}.xlsx"
                
                st.success("✅ План-График успешно сформирован!")
                
                st.download_button(
                    label="📥 Скачать готовый план",
                    data=excel_data,
                    file_name=file_title,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"❌ Ошибка при обработке данных: {str(e)}")
