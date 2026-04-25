import streamlit as st
import pandas as pd
import plotly.express as px
import io
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Рост отступлений", layout="wide")

# ==============================
# НОРМАЛИЗАЦИЯ (Защищенная)
# ==============================
def normalize_columns(df):
    """Приводит заголовки к единому стандарту и чистит их."""
    new_cols = {}
    for col in df.columns:
        c = str(col).strip().upper()  # str() защищает от нестроковых заголовков
        if c in ["КМ", "KM"]: new_cols[col] = "KM"
        elif c in ["М", "M"]: new_cols[col] = "M"
        elif c in ["КОДНАПРВ", "КОД"]: new_cols[col] = "KOD"
        elif c in ["ПУТЬ"]: new_cols[col] = "PATH"
        elif c in ["АМПЛИТУДА"]: new_cols[col] = "AMP"
        elif c in ["ДЛИНА"]: new_cols[col] = "LEN"
        elif c in ["СТЕПЕНЬ"]: new_cols[col] = "STEP"
        elif c in ["ОТСТУПЛЕНИЕ"]: new_cols[col] = "OTST"
        elif c in ["БАЛЛ"]: new_cols[col] = "BALL"
        elif c in ["ИС"]: new_cols[col] = "IS"
        elif c in ["СТРЕЛКА"]: new_cols[col] = "STR"
        elif c in ["МОСТ"]: new_cols[col] = "MOST"
        elif c in ["ГОД"]: new_cols[col] = "YEAR"
        elif c in ["МЕСЯЦ"]: new_cols[col] = "MONTH"
        elif c in ["ДЕНЬ"]: new_cols[col] = "DAY"
        else: new_cols[col] = col
    return df.rename(columns=new_cols)

def to_numeric_safe(series):
    """Безопасное преобразование в число: сначала в текст, потом замена запятой."""
    return pd.to_numeric(series.astype(str).str.replace(",", "."), errors="coerce")

def make_date_safe(df):
    """Собирает дату, если есть колонки Г/М/Д."""
    date_cols = ["YEAR", "MONTH", "DAY"]
    if all(col in df.columns for col in date_cols):
        return pd.to_datetime(
            df["YEAR"].astype(str) + "-" + df["MONTH"].astype(str) + "-" + df["DAY"].astype(str),
            errors="coerce"
        )
    return None

# ==============================
# ИНТЕРФЕЙС
# ==============================
st.title("📊 Анализ роста амплитуд")
st.write("Сравнение динамики состояния пути между двумя проверками.")

with st.sidebar:
    st.header("Настройки")
    file1 = st.file_uploader("📂 Файл №1 (Excel)", type=["xlsx"], key="f1")
    file2 = st.file_uploader("📂 Файл №2 (Excel)", type=["xlsx"], key="f2")
    tolerance = st.slider("Допуск по метрам (м)", 0, 10, 3)
    run_btn = st.button("🚀 Запустить анализ", use_container_width=True)

# ==============================
# ЛОГИКА ОБРАБОТКИ
# ==============================
if run_btn:
    if not file1 or not file2:
        st.error("Пожалуйста, загрузите оба файла Excel.")
        st.stop()

    try:
        # Чтение листов "Отступления"
        df1_raw = pd.read_excel(file1, sheet_name="Отступления")
        df2_raw = pd.read_excel(file2, sheet_name="Отступления")

        # Нормализация
        df1 = normalize_columns(df1_raw)
        df2 = normalize_columns(df2_raw)

        # Обработка дат для определения хронологии
        df1["DATE_STR"] = make_date_safe(df1)
        df2["DATE_STR"] = make_date_safe(df2)

        # Если даты есть, определяем кто старый, кто новый. Если нет — по порядку загрузки.
        if df1["DATE_STR"].notna().any() and df2["DATE_STR"].notna().any():
            if df1["DATE_STR"].min() < df2["DATE_STR"].min():
                old, new = df1, df2
            else:
                old, new = df2, df1
        else:
            old, new = df1, df2
            st.info("Даты в файлах не найдены. Сравнение идет в порядке загрузки (Файл 1 -> Файл 2).")

        # Подготовка данных (типы и очистка)
        for df in [old, new]:
            # Код и Путь — в текст без пробелов
            for col in ["KOD", "PATH", "OTST"]:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.strip()
            
            # Остальное — в числа
            for col in ["KM", "M", "AMP", "LEN", "BALL"]:
                if col in df.columns:
                    df[col] = to_numeric_safe(df[col])
        
        old = old.dropna(subset=["KM", "M", "AMP"])
        new = new.dropna(subset=["KM", "M", "AMP"])

        # Объединение (Merge)
        merged = pd.merge(
            old, new, 
            on=["KOD", "PATH", "KM", "OTST"], 
            suffixes=("_old", "_new")
        )

        # Фильтр по допуску метров
        merged["delta_m"] = abs(merged["M_new"] - merged["M_old"])
        merged = merged[merged["delta_m"] <= tolerance]

        # Убираем дубликаты (оставляем ближайшее по метрам)
        merged = merged.sort_values("delta_m").drop_duplicates(
            subset=["KOD", "PATH", "KM", "M_old", "OTST"], keep="first"
        )

        # Находим только РОСТ
        result_data = merged[merged["AMP_new"] > merged["AMP_old"]].copy()
        result_data["Рост"] = (result_data["AMP_new"] - result_data["AMP_old"]).round(1)

        if not result_data.empty:
            # Сборка финальной таблицы
            def safe_col(df, name): return df[name] if name in df.columns else ""

            df_final = pd.DataFrame({
                "Код": result_data["KOD"],
                "Путь": result_data["PATH"],
                "КМ": result_data["KM"],
                "М_было": result_data["M_old"],
                "М_стало": result_data["M_new"],
                "Тип": result_data["OTST"],
                "Амп_было": result_data["AMP_old"],
                "Амп_стало": result_data["AMP_new"],
                "Рост": result_data["Рост"],
                "Степень": safe_col(result_data, "STEP_new")
            })

            # --- ВИЗУАЛИЗАЦИЯ ---
            st.subheader("📈 Визуализация динамики")
            
            # Метрики
            m1, m2, m3 = st.columns(3)
            m1.metric("Всего точек роста", len(df_final))
            m2.metric("Макс. рост", f"{df_final['Рост'].max()} мм")
            m3.metric("Средний рост", f"{df_final['Рост'].mean():.1f} мм")

            # График
            fig = px.scatter(
                df_final, 
                x="КМ", 
                y="Рост", 
                size="Амп_стало", 
                color="Тип",
                hover_data=["М_стало", "Амп_было", "Амп_стало"],
                title="Распределение роста амплитуд по участку (размер = текущая амплитуда)",
                template="plotly_white"
            )
            st.plotly_chart(fig, use_container_width=True)

            # Таблица
            st.subheader("📋 Детальный список")
            st.dataframe(df_final.sort_values("Рост", ascending=False), use_container_width=True)

            # --- ЭКСПОРТ ---
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_final.to_excel(writer, index=False, sheet_name="Рост")
                ws = writer.book["Рост"]
                red = PatternFill(start_color="FF9999", fill_type="solid")
                # Красим столбец "Рост" (9-й по счету)
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=9).value > 5:
                        ws.cell(row=row, column=9).fill = red

            st.download_button(
                label="📥 Скачать отчет (Excel)",
                data=output.getvalue(),
                file_name="growth_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("Роста амплитуд в сопоставимых точках не обнаружено.")

    except Exception as e:
        st.error(f"Произошла ошибка: {e}")
